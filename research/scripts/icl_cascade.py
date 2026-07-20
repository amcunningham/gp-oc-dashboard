#!/usr/bin/env python3
"""
icl_cascade.py — reproducible analysis behind the "where the inverse care law
operates" Substack post and its supplement (research/icl_supplement.html).

Regenerates every number in the post and supplement from the source files:

  GPPS 2026 analysis-tool crosstab exports (weighted), in data/:
    GPPS_National_Crosstab_13072026.xlsx      Q8..Q32 x deprivation x Q41 (no filter)
    GPPS_National_Crosstab_11072026 (6).xlsx  Q12 x deprivation x Q41 (all ethnic groups)
    GPPS_National_Crosstab_13072026 (5).xlsx  Q12 x deprivation x Q9 reason (no filter)
    GPPS_National_Crosstab_11072026 (9).xlsx  Q12 x Q39 condition x deprivation
    GPPS_National_Crosstab_14072026.xlsx      all last-contact Qs x Q8 recency x Q41, most deprived
    GPPS_National_Crosstab_14072026 (1).xlsx  same, least deprived
    GPPS_National_Crosstab_14072026 (3).xlsx  all last-appointment Qs x Q17 recency x Q41, most deprived
    GPPS_National_Crosstab_14072026 (2).xlsx  same, least deprived
    GPPS_2026_National_results_and_trends_PUBLIC.xlsx  (Q12 national trend 2024-26)

  Practice-level panels, in research/data/:
    xsec_master_2026.parquet   (GPPS practice deflection 2025/2026, IMD, list size)
    panel_merged.parquet       (GPAD appointments + OC submissions, practice-month)
    cbt_volumes_panel.csv      (cloud-based telephony, practice-month, Oct 2024-May 2026)
    practice_age_sex.parquet   (age structure + registered patients, Jul 2026)
    qof_prevalence_2425.parquet (QOF recorded prevalence 2024/25)

  Workforce, in data/gpw_may26/:
    3 General Practice – May 2026 Practice Level - High level.csv
      (NHSE General Practice Workforce, May 2026, long format; GP FTE by role)

NOTE on deprivation direction: xsec_master_2026.imd_quintile runs 1 = LEAST
deprived to 5 = MOST deprived (quintiles of rising IMD score). The GPPS
analysis-tool exports run the other way (1 = most deprived). This script
handles both and always labels output "most_deprived"/"least_deprived".

Outputs:
    research/icl_cascade_results.json  (machine-readable, every figure keyed)
    research/icl_cascade_results.md    (human-readable, figure -> source -> value)

Run from anywhere inside the repo:  python research/scripts/icl_cascade.py
Percentages are recomputed from weighted counts where the tool gives them, so
they carry one decimal rather than the tool's on-screen rounding.
"""

from __future__ import annotations

import json
import math
import re
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
RDATA = ROOT / "research" / "data"
OUT_JSON = ROOT / "research" / "icl_cascade_results.json"
OUT_MD = ROOT / "research" / "icl_cascade_results.md"

FILES = {
    "dep_x_q41": DATA / "GPPS_National_Crosstab_13072026.xlsx",
    "q12_dep_x_q41": DATA / "GPPS_National_Crosstab_11072026 (6).xlsx",
    "q12_dep_x_reason": DATA / "GPPS_National_Crosstab_13072026 (5).xlsx",
    "q12_cond_x_dep": DATA / "GPPS_National_Crosstab_11072026 (9).xlsx",
    "dep_x_age": DATA / "GPPS_National_Crosstab_13072026 (1).xlsx",
    "mostdep_q8_x_q41": DATA / "GPPS_National_Crosstab_14072026.xlsx",
    "leastdep_q8_x_q41": DATA / "GPPS_National_Crosstab_14072026 (1).xlsx",
    "mostdep_q17_x_q41": DATA / "GPPS_National_Crosstab_14072026 (3).xlsx",
    "leastdep_q17_x_q41": DATA / "GPPS_National_Crosstab_14072026 (2).xlsx",
    "national_trends": DATA / "GPPS_2026_National_results_and_trends_PUBLIC.xlsx",
    "workforce_may26": DATA / "gpw_may26" /
        "3 General Practice – May 2026 Practice Level - High level.csv",
}

# xsec_master_2026 quintiles: 5 = most deprived, 1 = least deprived
XSEC_MOST, XSEC_LEAST = 5, 1


# ----------------------------------------------------------------------------
# Generic parser for analysis-tool crosstab exports
# ----------------------------------------------------------------------------

def parse_crosstab(path: Path) -> dict:
    """Parse an analysis-tool export into {meta, blocks}.

    Layout: 'Results' sheet; each question block starts with a 'Qnn.' title row,
    followed by a comparator-1 header row (groups spanning columns), an optional
    comparator-2 header row (subgroups), a 'Base' row, then answer rows of
    (proportion, weighted count) column pairs, ending at 'Unweighted Base'.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb["Results"].iter_rows(values_only=True))
    wb.close()

    meta = {}
    for r in rows[:8]:
        c0 = str(r[0]) if r[0] else ""
        for key, tag in (("Filters applied:", "filters"),
                         ("Comparator 1:", "comparator1"),
                         ("Comparator 2:", "comparator2")):
            if c0.startswith(key):
                meta[tag] = c0[len(key):].strip()

    qtitle = re.compile(r"^Q\d+\. ")
    blocks = []
    i = 0
    while i < len(rows):
        c0 = str(rows[i][0]) if rows[i][0] else ""
        if not qtitle.match(c0):
            i += 1
            continue
        title = c0
        # header rows: comparator1 groups, then (optionally) comparator2 subgroups,
        # then 'Base'
        hdr = i + 1
        header_rows = []
        while hdr < len(rows) and str(rows[hdr][0]) != "Base":
            header_rows.append(rows[hdr])
            hdr += 1
        base_row = rows[hdr]
        grp_row = header_rows[0]          # comparator-1 categories (spanning)
        sub_row = header_rows[-1] if len(header_rows) > 1 else None

        # map columns: col -> (group, subgroup); value cols come in (prop, count) pairs
        ncol = len(base_row)
        groups = {}
        cur = None
        for c in range(1, ncol):
            g = grp_row[c]
            if g is not None and str(g).strip():
                cur = str(g).strip()
            groups[c] = cur
        subs = {}
        cur = None
        for c in range(1, ncol):
            if sub_row is not None:
                s = sub_row[c]
                if s is not None and str(s).strip():
                    cur = str(s).strip()
            subs[c] = cur if sub_row is not None else None

        # answer rows until 'Unweighted Base'
        j = hdr + 1
        answers = {}
        bases = {}
        for c in range(1, ncol, 2):
            key = (groups[c], subs[c])
            if base_row[c] is not None:
                bases[key] = float(base_row[c])
        while j < len(rows):
            lab = str(rows[j][0]) if rows[j][0] else ""
            if lab.startswith("Unweighted Base") or (qtitle.match(lab) and j > hdr + 1):
                break
            if lab:
                vals = {}
                for c in range(1, ncol, 2):
                    key = (groups[c], subs[c])
                    prop = rows[j][c]
                    cnt = rows[j][c + 1] if c + 1 < ncol else None
                    if prop is not None or cnt is not None:
                        vals[key] = (None if prop is None else float(prop),
                                     None if cnt is None else float(cnt))
                answers.setdefault(lab, vals)
            j += 1
        blocks.append({"title": title, "bases": bases, "answers": answers})
        i = j
    return {"meta": meta, "blocks": blocks}


def block(ct: dict, qprefix: str) -> dict:
    for b in ct["blocks"]:
        if b["title"].startswith(qprefix):
            return b
    raise KeyError(f"no block starting {qprefix!r}")


def rate(b: dict, answer_label: str, group: str, sub: str,
         answer_labels: list[str] | None = None) -> dict:
    """Percentage (from weighted counts / base), base, and binomial 95% CI.

    If answer_labels is given, the numerator is the sum of those rows
    (used for the diversion measure, which sums three Q14 options).
    """
    key = (group, sub)
    n = b["bases"][key]
    labels = answer_labels or [answer_label]
    k = 0.0
    for lab in labels:
        # allow prefix match (labels are sometimes truncated in exports)
        matches = [a for a in b["answers"] if a.startswith(lab)]
        if not matches:
            raise KeyError(f"answer {lab!r} not found")
        v = b["answers"][matches[0]].get(key)
        if v is None or v[1] is None:
            # fall back to proportion x base
            k += (v[0] or 0.0) * n
        else:
            k += v[1]
    p = k / n
    se = math.sqrt(p * (1 - p) / n)
    return {"pct": round(100 * p, 1), "n": round(n),
            "ci95": [round(100 * (p - 1.96 * se), 1), round(100 * (p + 1.96 * se), 1)]}


def gap(a: dict, b_: dict) -> dict:
    """Difference in percentage points with a 95% CI (independent binomials)."""
    p1, n1 = a["pct"] / 100, a["n"]
    p2, n2 = b_["pct"] / 100, b_["n"]
    d = p1 - p2
    se = math.sqrt(p1 * (1 - p1) / n1 + p2 * (1 - p2) / n2)
    return {"pp": round(100 * d, 1),
            "ci95": [round(100 * (d - 1.96 * se), 1), round(100 * (d + 1.96 * se), 1)]}


DIVERSION_LABELS = ["Told to go to a pharmacy",
                    "Told to contact NHS 111 or a different NHS s",
                    "Told to get urgent care"]
COMEBACK = "I was told to contact my practice again anot"
LAST3M = "In the last 3 months"
ALOT, ALITTLE, NONE_ = "Yes, a lot", "Yes, a little", "No, not at all"
QUINTILES = ["1 (Most deprived)", "2", "3", "4", "5 (Least deprived)"]


def main() -> None:
    res: dict = {"_provenance": {k: str(v.relative_to(ROOT)) for k, v in FILES.items()}}

    # ------------------------------------------------------------------ GPPS
    dep41 = parse_crosstab(FILES["dep_x_q41"])
    md_q8 = parse_crosstab(FILES["mostdep_q8_x_q41"])
    ld_q8 = parse_crosstab(FILES["leastdep_q8_x_q41"])
    md_q17 = parse_crosstab(FILES["mostdep_q17_x_q41"])
    ld_q17 = parse_crosstab(FILES["leastdep_q17_x_q41"])

    # 1. Need by quintile: Q41 weighted bases (from the Q8 block's Base row)
    q8b = block(dep41, "Q8.")
    res["need_by_quintile"] = {
        q: {lim: round(q8b["bases"][(q, lim)]) for lim in (ALOT, ALITTLE, NONE_)}
        for q in QUINTILES
    }

    # 2. Presentation (Q8 recency) among the most affected, by quintile
    pres = {}
    for q in QUINTILES:
        last3 = rate(q8b, LAST3M, q, ALOT)
        last6 = rate(q8b, "Summary Statistic - % In the last 6 mont", q, ALOT)
        no12 = rate(q8b, "More than 12 months ago", q, ALOT,
                    ["More than 12 months ago", "I haven’t contacted my GP practice since"])
        pres[q] = {"last_3m": last3, "within_6m": last6,
                   "not_in_3m": {"pct": round(100 - last3["pct"], 1), "n": last3["n"],
                                 "ci95": [round(100 - last3["ci95"][1], 1),
                                          round(100 - last3["ci95"][0], 1)]},
                   "not_in_12m_or_never": no12}
    res["presentation_q8_alot"] = pres

    # 3. The 3-month cascade, most affected, most vs least deprived
    #    Rows 2-3 come from the quintile-filtered Q8-comparator files (last
    #    contact in the last 3 months); row 4 from the Q17-comparator files
    #    (last appointment in the last 3 months, Q31 excl. don't know).
    def cascade_row(md_ct, ld_ct, qprefix, label, sublabels=None):
        m = rate(block(md_ct, qprefix), label, LAST3M, ALOT, sublabels)
        l = rate(block(ld_ct, qprefix), label, LAST3M, ALOT, sublabels)
        return {"most_deprived": m, "least_deprived": l, "gap": gap(m, l)}

    pres_gap = gap(pres["1 (Most deprived)"]["not_in_3m"],
                   pres["5 (Least deprived)"]["not_in_3m"])
    res["cascade_3m_alot"] = {
        "didnt_contact_3m": {"most_deprived": pres["1 (Most deprived)"]["not_in_3m"],
                             "least_deprived": pres["5 (Least deprived)"]["not_in_3m"],
                             "gap": pres_gap},
        "comeback": cascade_row(md_q8, ld_q8, "Q12.", COMEBACK),
        "diversion": cascade_row(md_q8, ld_q8, "Q14.", None, DIVERSION_LABELS),
        "needs_not_met": cascade_row(md_q17, ld_q17, "Q31.", "No, not at all"),
    }

    # 4. The 12-month/6-month variant used in earlier drafts (reconciliation)
    res["cascade_earlier_variant"] = {
        "didnt_contact_12m_or_never": {
            "most_deprived": pres["1 (Most deprived)"]["not_in_12m_or_never"],
            "least_deprived": pres["5 (Least deprived)"]["not_in_12m_or_never"]},
        "needs_not_met_6m_contact": {
            "most_deprived": rate(block(md_q8, "Q31."), "No, not at all",
                                  "Summary Statistic - % In the last 6 months", ALOT),
            "least_deprived": rate(block(ld_q8, "Q31."), "No, not at all",
                                   "Summary Statistic - % In the last 6 months", ALOT)},
    }

    # 5. Come-back by limitation (3-month contacts), most vs least deprived
    res["comeback_by_limitation_3m"] = {
        dep: {lim: rate(block(ct, "Q12."), COMEBACK, LAST3M, lim)
              for lim in (ALOT, ALITTLE, NONE_)}
        for dep, ct in (("most_deprived", md_q8), ("least_deprived", ld_q8))
    }

    # 6. Diversion by limitation (3-month contacts)
    res["diversion_by_limitation_3m"] = {
        dep: {lim: rate(block(ct, "Q14."), None, LAST3M, lim, DIVERSION_LABELS)
              for lim in (ALOT, ALITTLE, NONE_)}
        for dep, ct in (("most_deprived", md_q8), ("least_deprived", ld_q8))
    }

    # 7. Needs-not-met by limitation (appointment in last 3 months)
    res["needs_not_met_by_limitation_3m"] = {
        dep: {lim: rate(block(ct, "Q31."), "No, not at all", LAST3M, lim)
              for lim in (ALOT, ALITTLE, NONE_)}
        for dep, ct in (("most_deprived", md_q17), ("least_deprived", ld_q17))
    }

    # 8. Come-back by reason for contact (Q9), most vs least deprived (all patients)
    reason_ct = parse_crosstab(FILES["q12_dep_x_reason"])
    q12r = block(reason_ct, "Q12.")
    reasons = sorted({s for (_, s) in q12r["bases"] if s})
    res["comeback_by_reason"] = {
        dep_label: {s: rate(q12r, COMEBACK, q, s) for s in reasons}
        for dep_label, q in (("most_deprived", "1 (Most deprived)"),
                             ("least_deprived", "5 (Least deprived)"))
    }

    # 9. Come-back by long-term condition (Q39) x deprivation
    cond_ct = parse_crosstab(FILES["q12_cond_x_dep"])
    q12c = block(cond_ct, "Q12.")
    conditions = sorted({g for (g, _) in q12c["bases"] if g and g != "Total"})
    res["comeback_by_condition"] = {
        cond: {dep_label: rate(q12c, COMEBACK, cond, q)
               for dep_label, q in (("most_deprived", "1 (Most deprived)"),
                                    ("least_deprived", "5 (Least deprived)"))}
        for cond in conditions
    }

    # 10. Come-back x deprivation x limitation (12-month base file, all contacts)
    v6 = parse_crosstab(FILES["q12_dep_x_q41"])
    q12v = block(v6, "Q12.")
    res["comeback_by_quintile_alot_allcontacts"] = {
        q: rate(q12v, COMEBACK, q, ALOT) for q in QUINTILES
    }

    # 10b. Age robustness: come-back and diversion gaps within every age band
    age_ct = parse_crosstab(FILES["dep_x_age"])
    q12a, q14a = block(age_ct, "Q12."), block(age_ct, "Q14.")
    age_bands = [s for (g, s) in q12a["bases"]
                 if g == "1 (Most deprived)" and s and not s.startswith("Summary")]
    rob = {}
    for band in age_bands:
        cb_m = rate(q12a, COMEBACK, "1 (Most deprived)", band)
        cb_l = rate(q12a, COMEBACK, "5 (Least deprived)", band)
        dv_m = rate(q14a, None, "1 (Most deprived)", band, DIVERSION_LABELS)
        dv_l = rate(q14a, None, "5 (Least deprived)", band, DIVERSION_LABELS)
        rob[band] = {"comeback_gap_pp": gap(cb_m, cb_l)["pp"],
                     "comeback": [cb_m["pct"], cb_l["pct"]],
                     "diversion_gap_pp": gap(dv_m, dv_l)["pp"],
                     "diversion": [dv_m["pct"], dv_l["pct"]]}
    res["age_band_robustness"] = rob

    # 11. National come-back trend 2024-2026
    wb = openpyxl.load_workbook(FILES["national_trends"], read_only=True)
    ws = wb["National results and trends"]
    trend = None
    for r in ws.iter_rows(values_only=True):
        if r[0] and str(r[0]).startswith("I was told to contact my practice again"):
            trend = [round(100 * float(v), 1) for v in r[1:4]]
    wb.close()
    res["national_comeback_trend"] = dict(zip(["2024", "2025", "2026"], trend))

    # ------------------------------------------------- practice-level (xsec)
    x = pd.read_parquet(RDATA / "xsec_master_2026.parquet")

    def wmean(df, col, w="list_size"):
        d = df.dropna(subset=[col, w])
        return (d[col] * d[w]).sum() / d[w].sum()

    # deflection_20xx is stored as a percentage (e.g. 12.6 = 12.6%)
    res["practice_deflection_by_quintile"] = {
        yr: {("most_deprived" if q == XSEC_MOST else
              "least_deprived" if q == XSEC_LEAST else f"quintile_{int(q)}"):
             round(wmean(x[x.imd_quintile == q], f"deflection_{yr}"), 1)
             for q in sorted(x.imd_quintile.dropna().unique())}
        for yr in (2025, 2026)
    }
    res["n_practices_xsec"] = int(len(x))

    # --------------------------------------------- workforce (NHSE, May 2026)
    # Long format: one row per practice x staff role x measure. "Total" GP FTE
    # includes registrars and locums; fully-qualified = Total minus the five
    # "GP in Training Grade" roles. Denominator: registered patients from the
    # NHSE "Patients Registered at a GP Practice" snapshot (practice_age_sex,
    # Jul 2026) — the repo's canonical list-size source. Rates are ratios of
    # sums (total FTE / total patients), i.e. list-size-weighted.
    gpw = pd.read_csv(FILES["workforce_may26"])
    gpfte = gpw[(gpw.STAFF_GROUP == "GP") & (gpw.MEASURE == "FTE")]
    piv = gpfte.pivot_table(index="PRAC_CODE", columns="DETAILED_STAFF_ROLE",
                            values="VALUE", aggfunc="sum")
    training = [c for c in piv.columns if c.startswith("GP in Training")]
    piv["total_gp_fte"] = piv["Total"]
    piv["qual_gp_fte"] = piv["Total"] - piv[training].sum(axis=1)
    ages = pd.read_parquet(RDATA / "practice_age_sex.parquet")
    wf = (piv.reset_index().rename(columns={"PRAC_CODE": "gp_code"})
          .merge(ages[["gp_code", "total_list"]], on="gp_code")
          .merge(x[["gp_code", "imd_quintile"]], on="gp_code")
          .dropna(subset=["imd_quintile", "total_list"]))
    res["gp_fte_per_10k_may26"] = {}
    for label, q in (("most_deprived", XSEC_MOST), ("least_deprived", XSEC_LEAST)):
        d = wf[wf.imd_quintile == q]
        res["gp_fte_per_10k_may26"][label] = {
            "fully_qualified": round(d.qual_gp_fte.sum() / d.total_list.sum() * 10_000, 2),
            "total_incl_trainees_locums":
                round(d.total_gp_fte.sum() / d.total_list.sum() * 10_000, 2),
            "n_practices": int(len(d)),
        }
    mq = res["gp_fte_per_10k_may26"]["most_deprived"]
    lq = res["gp_fte_per_10k_may26"]["least_deprived"]
    res["gp_fte_per_10k_may26"]["pct_fewer_in_most_deprived"] = {
        "fully_qualified": round(100 * (1 - mq["fully_qualified"] / lq["fully_qualified"])),
        "total_incl_trainees_locums":
            round(100 * (1 - mq["total_incl_trainees_locums"] / lq["total_incl_trainees_locums"])),
    }

    # ------------------------------------------- admin demand/capacity table
    # Telephony x GPAD/OC, Oct 2024 - Mar 2026, practices in both panels.
    # Rates are ratios of sums: total volume / total registered patient-months
    # x 1,000 (equivalently, list-size-weighted practice-month means).
    cbt = pd.read_csv(RDATA / "cbt_volumes_panel.csv")
    cbt = cbt[(cbt.month >= "2024-10") & (cbt.month <= "2026-03")]
    pm = pd.read_parquet(RDATA / "panel_merged.parquet")
    pm = pm[(pm.month >= "2024-10") & (pm.month <= "2026-03")]
    m = cbt.merge(pm, on=["gp_code", "month"], how="inner")
    m = m.dropna(subset=["list_size", "imd_quintile", "inbound_calls",
                         "answered_calls", "total"])
    res["n_practices_telephony"] = int(m.gp_code.nunique())

    ages["pct65plus"] = (ages.a65_74 + ages.a75_84 + ages.a85plus) / ages.total_list * 100
    qof = pd.read_parquet(RDATA / "qof_prevalence_2425.parquet")[["gp_code", "dm_prev"]]
    prac = (m.groupby(["gp_code", "imd_quintile"])["list_size"].mean().reset_index()
            .merge(ages[["gp_code", "pct65plus"]], on="gp_code", how="left")
            .merge(qof, on="gp_code", how="left"))

    admin = {}
    for label, q in (("most_deprived", XSEC_MOST), ("least_deprived", XSEC_LEAST)):
        d = m[m.imd_quintile == q]
        pmonths = d.list_size.sum()  # registered patient-months
        oc = d.oc_total.fillna(0).sum()
        admin[label] = {
            "inbound_calls_per_1k_month": round(d.inbound_calls.sum() / pmonths * 1000),
            "answered_calls_per_1k_month": round(d.answered_calls.sum() / pmonths * 1000),
            "oc_submissions_per_1k_month": round(oc / pmonths * 1000),
            "patient_initiated_contacts_per_1k_month":
                round((d.answered_calls.sum() + oc) / pmonths * 1000),
            "appointments_per_1k_month": round(d.total.sum() / pmonths * 1000),
            "pct_aged_65_plus": round(wmean(prac[prac.imd_quintile == q], "pct65plus"), 1),
            "diabetes_prevalence_pct":
                round(wmean(prac[prac.imd_quintile == q], "dm_prev"), 1),
            "n_practices": int(d.gp_code.nunique()),
        }
    res["admin_demand_capacity"] = admin

    # telephony coverage: share of xsec practices present in the merged sample
    tele = set(m.gp_code)
    res["telephony_coverage_pct"] = {
        label: round(100 * len(tele & set(x[x.imd_quintile == q].gp_code)) /
                     len(x[x.imd_quintile == q]))
        for label, q in (("most_deprived", XSEC_MOST), ("least_deprived", XSEC_LEAST))
    }

    # ------------------------------------------------------------- outputs
    OUT_JSON.write_text(json.dumps(res, indent=1))

    def fmt(v, depth=0):
        pad = "  " * depth
        if isinstance(v, dict):
            if set(v) == {"pct", "n", "ci95"}:
                return f"{v['pct']}% (95% CI {v['ci95'][0]}–{v['ci95'][1]}; n={v['n']:,})"
            if set(v) == {"pp", "ci95"}:
                return f"{v['pp']}pp (95% CI {v['ci95'][0]}–{v['ci95'][1]})"
            return "\n" + "\n".join(f"{pad}- **{k}**: {fmt(w, depth + 1)}"
                                    for k, w in v.items())
        return str(v)

    lines = ["# icl_cascade results",
             "",
             "Every figure in the post and supplement, regenerated from source.",
             "Regenerate with `python research/scripts/icl_cascade.py`.", ""]
    for k, v in res.items():
        lines.append(f"## {k}")
        lines.append(fmt(v))
        lines.append("")
    OUT_MD.write_text("\n".join(lines))
    print(f"wrote {OUT_JSON} and {OUT_MD}")


if __name__ == "__main__":
    main()
